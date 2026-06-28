# -*- coding: utf-8 -*-
"""
모듈/스테이지 엑셀(.xlsx)을 게임이 읽는 JSON 으로 변환한다.
엑셀을 편집한 뒤 이 스크립트를 실행하면 Assets/Resources/Tables/*.json 이 갱신된다.
"""
import os
import json
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TABLES_DIR = os.path.join(ROOT, "Assets", "Tables")
OUT_DIR = os.path.join(ROOT, "Assets", "Resources", "Tables")
os.makedirs(OUT_DIR, exist_ok=True)


def read_rows(xlsx_path):
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    result = []
    for raw in rows[1:]:
        if raw[0] is None:
            continue
        result.append(dict(zip(headers, raw)))
    return result


def write_json(name, payload):
    path = os.path.join(OUT_DIR, name)
    with open(path, "w", encoding="utf-8") as stream:
        json.dump(payload, stream, ensure_ascii=False, indent=2)
    print("wrote:", path)


def convert_modules():
    rows = read_rows(os.path.join(TABLES_DIR, "Modules.xlsx"))
    out = []
    for row in rows:
        out.append({
            "Type": str(row["Type"]),
            "Category": str(row["Category"]),
            "DisplayName": str(row["DisplayName"]),
            "Price": int(row["Price"]),
            "Attack": int(row["Attack"]),
            "Health": int(row["Health"]),
            "Armor": int(row["Armor"]),
            "Speed": int(row["Speed"]),
            "Range": int(row["Range"]),
            "PowerSupply": int(row["PowerSupply"]),
            "PowerCost": int(row["PowerCost"]),
            "Color": str(row["ColorHex"]),
        })
    write_json("Modules.json", {"rows": out})


def convert_stages():
    rows = read_rows(os.path.join(TABLES_DIR, "Stages.xlsx"))
    out = []
    for row in rows:
        out.append({
            "Index": int(row["Index"]),
            "EnemyName": str(row["EnemyName"]),
            "ModuleCount": int(row["ModuleCount"]),
            "MaxTier": int(row["MaxTier"]),
            "Seed": int(row["Seed"]),
        })
    write_json("Stages.json", {"rows": out})


if __name__ == "__main__":
    convert_modules()
    convert_stages()
